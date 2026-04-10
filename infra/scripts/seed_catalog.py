"""
Seed the AI Governance Platform database from the Excel dataset.

Loads: data/UN_AI_Governance_Dataset_v2.xlsx (or DATASET_PATH override)
Seeds:
  1. Controls sheet                -> control table
  2. Requirements (Flat)           -> regulation + requirement tables
  3. Control Metrics sheet         -> control_metric_definition table
  4. Requirements by Control       -> control_requirement join table
  5. Baseline risk definitions     -> risk_definition table
  6. Interpretation Versioning     -> risk_interpretation table

Run:
  python infra/scripts/seed_catalog.py

Env:
  DATABASE_URL  (default: postgresql+asyncpg://aigov:localdev@localhost:5432/aigov)
  DATASET_PATH  (optional absolute/relative .xlsx path)

Notes:
  - Idempotent for all unique-constrained entities.
  - Requirements are upserted by code, so previously placeholder-linked rows are corrected.
  - Interpretations are de-duplicated by requirement + layer + content.
"""
from __future__ import annotations

import asyncio
import json
import os
import uuid
from pathlib import Path

import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine

DEFAULT_DATASET_PATH = Path(__file__).parents[2] / "data" / "UN_AI_Governance_Dataset_v2.xlsx"
DATASET_PATH = Path(os.getenv("DATASET_PATH", str(DEFAULT_DATASET_PATH)))
DEFAULT_DB_URL = "postgresql+asyncpg://aigov:localdev@localhost:5432/aigov"

TIER_MAP = {
    "Foundation": "FOUNDATION",
    "Common": "COMMON",
    "Specialized": "SPECIALIZED",
}

PLACEHOLDER_REG_ID = "00000000-0000-0000-0000-000000000001"

RISK_DEFINITIONS = [
    {
        "code": "TIER_FOUNDATION",
        "description": "Baseline risk tier for low-impact AI systems.",
        "tier_floor": "Foundation",
    },
    {
        "code": "TIER_COMMON",
        "description": "Medium risk tier where additional controls apply.",
        "tier_floor": "Common",
    },
    {
        "code": "TIER_HIGH",
        "description": "High risk tier for sensitive domains and high-impact use cases.",
        "tier_floor": "High",
    },
    {
        "code": "FLOOR_ASYLUM",
        "description": "Asylum-related AI systems are floored to High tier.",
        "tier_floor": "High",
    },
    {
        "code": "FLOOR_CRIMINAL_JUSTICE",
        "description": "Criminal justice AI systems are floored to High tier.",
        "tier_floor": "High",
    },
    {
        "code": "FLOOR_MEDICAL_DIAGNOSIS",
        "description": "Medical diagnosis AI systems are floored to High tier.",
        "tier_floor": "High",
    },
    {
        "code": "FLOOR_BIOMETRIC_ID",
        "description": "Biometric identification AI systems are floored to High tier.",
        "tier_floor": "High",
    },
    {
        "code": "AUTONOMY_VALIDATION",
        "description": "If declared HITL conflicts with observed override behavior, elevate autonomy risk.",
        "tier_floor": "High",
    },
]


def _new_id() -> str:
    return str(uuid.uuid4())


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _load_workbook() -> openpyxl.Workbook:
    if not DATASET_PATH.exists():
        raise FileNotFoundError(f"Dataset not found: {DATASET_PATH}")
    return openpyxl.load_workbook(DATASET_PATH, read_only=True, data_only=True)


def _sheet_rows(ws) -> list[dict]:
    """Return all rows as dicts keyed by header, skipping the header row itself."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_clean(h) if h is not None else None for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def _normalize_layer(layer_raw: str) -> str | None:
    low = layer_raw.lower()
    if "source" in low or "layer 1" in low:
        return "SOURCE"
    if "system" in low or "layer 2" in low:
        return "SYSTEM"
    if "user" in low or "layer 3" in low:
        return "USER"
    return None
GOVERNANCE_CATEGORY_ALIAS = {
    "all": "Risk & Compliance",
    "risk management": "Risk & Compliance",
    "regulatory": "Risk & Compliance",
    "privacy": "Risk & Compliance",
    "governance": "Corporate Oversight",
    "security": "Security",
    "operation": "System Performance",
    "operations": "System Performance",
    "incident management": "System Performance",
    "development": "Technical Architecture",
    "deployment": "Infrastructure",
    "design": "Solution Design",
    "lifecycle": "Data Integration",
    "responsible systems": "Solution Design",
    "communication": "Corporate Oversight",
    "audit": "Risk & Compliance",
    "third party": "Risk & Compliance",
    "un-specific": "Corporate Oversight",
}


def _normalize_governance_category(*values: str) -> str:
    for value in values:
        key = _clean(value).lower()
        if not key:
            continue
        if key in GOVERNANCE_CATEGORY_ALIAS:
            return GOVERNANCE_CATEGORY_ALIAS[key]
    return "Risk & Compliance"


# ---------------------------------------------------------------------------
# Regulation helpers
# ---------------------------------------------------------------------------

async def _load_regulation_map(conn) -> dict[tuple[str, str], str]:
    result = await conn.execute(
        text(
            """
            SELECT id::text AS id, title, COALESCE(jurisdiction, '') AS jurisdiction
            FROM regulation
            """
        )
    )
    reg_map: dict[tuple[str, str], str] = {}
    for row in result.mappings().all():
        key = (_clean(row["title"]).lower(), _clean(row["jurisdiction"]).lower())
        reg_map[key] = row["id"]
    return reg_map


async def _ensure_placeholder_regulation(conn, reg_map: dict[tuple[str, str], str]) -> None:
    await conn.execute(
        text(
            """
            INSERT INTO regulation (id, title, jurisdiction)
            VALUES (:id, :title, :jurisdiction)
            ON CONFLICT (id) DO NOTHING
            """
        ),
        {
            "id": PLACEHOLDER_REG_ID,
            "title": "Unlinked - Phase 3",
            "jurisdiction": "PLACEHOLDER",
        },
    )

    placeholder_row = await conn.execute(
        text(
            """
            SELECT id::text AS id, title, COALESCE(jurisdiction, '') AS jurisdiction
            FROM regulation
            WHERE id = :id
            """
        ),
        {"id": PLACEHOLDER_REG_ID},
    )
    row = placeholder_row.mappings().first()
    if row is None:
        return

    # Preserve canonical lookup key even if historical title variants exist.
    reg_map[("unlinked - phase 3", "placeholder")] = row["id"]
    reg_map[(_clean(row["title"]).lower(), _clean(row["jurisdiction"]).lower())] = row["id"]


async def _get_or_create_regulation(
    conn,
    reg_map: dict[tuple[str, str], str],
    title: str,
    jurisdiction: str,
) -> tuple[str, bool]:
    normalized_title = title or "Unlinked - Phase 3"
    normalized_jurisdiction = jurisdiction or "PLACEHOLDER"
    key = (normalized_title.lower(), normalized_jurisdiction.lower())

    if key in reg_map:
        return reg_map[key], False

    new_id = _new_id()
    await conn.execute(
        text(
            """
            INSERT INTO regulation (id, title, jurisdiction)
            VALUES (:id, :title, :jurisdiction)
            """
        ),
        {
            "id": new_id,
            "title": normalized_title,
            "jurisdiction": normalized_jurisdiction,
        },
    )
    reg_map[key] = new_id
    return new_id, True


# ---------------------------------------------------------------------------
# Sheet loaders
# ---------------------------------------------------------------------------

async def seed_controls(conn) -> int:
    wb = _load_workbook()
    ws = wb["Controls"]
    rows = _sheet_rows(ws)
    count = 0

    for row in rows:
        code = _clean(row.get("Control ID"))
        if not code or code.startswith("--") or code.startswith("["):
            continue

        tier_raw = _clean(row.get("Tier"))
        tier = TIER_MAP.get(tier_raw, tier_raw.upper() if tier_raw else None)
        is_foundation = tier_raw == "Foundation"
        domain = _clean(row.get("Domain")).lower() or None

        mode_raw = _clean(row.get("Measurement Mode")).lower()
        measurement_mode = mode_raw if mode_raw in ("system_calculated", "hybrid", "manual") else None

        await conn.execute(
            text(
                """
                INSERT INTO control (id, code, title, description, domain, tier, is_foundation, measurement_mode)
                VALUES (:id, :code, :title, :description, :domain, :tier, :is_foundation, :measurement_mode)
                ON CONFLICT (code) DO UPDATE
                SET
                  title = EXCLUDED.title,
                  description = EXCLUDED.description,
                  domain = EXCLUDED.domain,
                  tier = EXCLUDED.tier,
                  is_foundation = EXCLUDED.is_foundation,
                  measurement_mode = EXCLUDED.measurement_mode
                """
            ),
            {
                "id": _new_id(),
                "code": code,
                "title": _clean(row.get("Control Name")) or code,
                "description": _clean(row.get("Description")) or None,
                "domain": domain,
                "tier": tier,
                "is_foundation": is_foundation,
                "measurement_mode": measurement_mode,
            },
        )
        count += 1

    print(f"  Controls processed: {count} rows")
    wb.close()
    return count


async def seed_requirements(conn) -> tuple[int, int]:
    wb = _load_workbook()
    ws = wb["Requirements (Flat)"]
    rows = _sheet_rows(ws)

    reg_map = await _load_regulation_map(conn)
    await _ensure_placeholder_regulation(conn, reg_map)

    req_count = 0
    new_reg_count = 0

    for row in rows:
        code = _clean(row.get("Requirement ID"))
        if not code:
            continue

        source = _clean(row.get("Source"))
        jurisdiction = _clean(row.get("Jurisdiction"))
        regulation_id, created = await _get_or_create_regulation(conn, reg_map, source, jurisdiction)
        if created:
            new_reg_count += 1

        lifecycle_stage = _clean(row.get("Lifecycle Stage"))
        source_ref = _clean(row.get("Source Ref"))

        await conn.execute(
            text(
                """
                INSERT INTO requirement (id, regulation_id, code, title, description, category)
                VALUES (:id, :regulation_id, :code, :title, :description, :category)
                ON CONFLICT (code) DO UPDATE
                SET
                  regulation_id = EXCLUDED.regulation_id,
                  title = EXCLUDED.title,
                  description = EXCLUDED.description,
                  category = EXCLUDED.category
                """
            ),
            {
                "id": _new_id(),
                "regulation_id": regulation_id,
                "code": code,
                "title": _clean(row.get("Requirement Text")) or code,
                "description": source_ref or None,
                "category": _normalize_governance_category(lifecycle_stage, source),
            },
        )
        req_count += 1

    print(f"  Requirements processed: {req_count} rows")
    print(f"  Regulations created from dataset metadata: {new_reg_count}")
    wb.close()
    return req_count, new_reg_count


async def seed_metric_definitions(conn) -> int:
    result = await conn.execute(text("SELECT id, code FROM control"))
    control_by_code = {row.code: row.id for row in result}

    wb = _load_workbook()
    ws = wb["Control Metrics"]
    rows = _sheet_rows(ws)
    count = 0

    for row in rows:
        control_code = _clean(row.get("Control ID"))
        metric_key = _clean(row.get("Metric Key (OTEL)"))
        if not control_code or not metric_key:
            continue

        control_id = control_by_code.get(control_code)
        if not control_id:
            print(f"  WARNING: control '{control_code}' not found - skipping metric '{metric_key}'")
            continue

        threshold = {
            "compliant": _clean(row.get("Threshold: Compliant")) or None,
            "warning": _clean(row.get("Threshold: Warning")) or None,
            "breach": _clean(row.get("Threshold: Breach")) or None,
            "direction": _clean(row.get("Direction")) or None,
            "unit": _clean(row.get("Unit")) or None,
            "source_system": _clean(row.get("Source System")) or None,
            "calculation_type": _clean(row.get("Calculation Type")) or None,
            "formula": _clean(row.get("Formula / Logic")) or None,
            "delta_period": _clean(row.get("Delta Period")) or None,
            "peer_group_enabled": _clean(row.get("Peer Group Enabled")) or None,
        }

        is_manual = _clean(row.get("Measurement Mode")).lower() == "manual"

        await conn.execute(
            text(
                """
                INSERT INTO control_metric_definition (id, control_id, metric_name, threshold, is_manual)
                VALUES (:id, :control_id, :metric_name, :threshold, :is_manual)
                ON CONFLICT (control_id, metric_name) DO UPDATE
                SET
                  threshold = EXCLUDED.threshold,
                  is_manual = EXCLUDED.is_manual
                """
            ),
            {
                "id": _new_id(),
                "control_id": control_id,
                "metric_name": metric_key,
                "threshold": json.dumps(threshold),
                "is_manual": is_manual,
            },
        )
        count += 1

    print(f"  Metric definitions processed: {count} rows")
    wb.close()
    return count


async def seed_control_requirements(conn) -> int:
    """
    Populate control_requirement from 'Requirements by Control'.
    Control ID column is sparse: appears once per group, then blank.
    """
    ctrl_result = await conn.execute(text("SELECT id, code FROM control"))
    control_by_code = {row.code: row.id for row in ctrl_result}

    req_result = await conn.execute(text("SELECT id, code FROM requirement"))
    requirement_by_code = {row.code: row.id for row in req_result}

    wb = _load_workbook()
    ws = wb["Requirements by Control"]
    rows = _sheet_rows(ws)

    count = 0
    skipped = 0
    current_ctrl = None

    for row in rows:
        ctrl_code = _clean(row.get("Control ID"))
        req_code = _clean(row.get("Requirement ID"))

        if ctrl_code:
            current_ctrl = ctrl_code

        if not req_code:
            continue

        if req_code.startswith("--") or (req_code.startswith("[") and req_code.endswith("]")):
            skipped += 1
            continue

        if not current_ctrl:
            skipped += 1
            continue

        control_id = control_by_code.get(current_ctrl)
        requirement_id = requirement_by_code.get(req_code)

        if not control_id or not requirement_id:
            skipped += 1
            continue

        await conn.execute(
            text(
                """
                INSERT INTO control_requirement (control_id, requirement_id)
                VALUES (:control_id, :requirement_id)
                ON CONFLICT DO NOTHING
                """
            ),
            {"control_id": control_id, "requirement_id": requirement_id},
        )
        count += 1

    print(f"  Control-requirement links: {count} inserted, {skipped} skipped")
    wb.close()
    return count


async def seed_risk_definitions(conn) -> int:
    count = 0
    for item in RISK_DEFINITIONS:
        await conn.execute(
            text(
                """
                INSERT INTO risk_definition (id, code, description, tier_floor)
                VALUES (:id, :code, :description, :tier_floor)
                ON CONFLICT (code) DO UPDATE
                SET
                  description = EXCLUDED.description,
                  tier_floor = EXCLUDED.tier_floor
                """
            ),
            {
                "id": _new_id(),
                "code": item["code"],
                "description": item["description"],
                "tier_floor": item["tier_floor"],
            },
        )
        count += 1

    print(f"  Risk definitions upserted: {count}")
    return count


async def seed_interpretations(conn) -> tuple[int, int]:
    req_result = await conn.execute(text("SELECT id, code FROM requirement"))
    requirement_by_code = {row.code: row.id for row in req_result}

    existing_rows = await conn.execute(
        text(
            """
            SELECT requirement_id::text AS requirement_id,
                   layer::text AS layer,
                   COALESCE(version, 0) AS version,
                   content
            FROM risk_interpretation
            """
        )
    )

    existing_keys: set[tuple[str, str, str]] = set()
    next_versions: dict[tuple[str, str], int] = {}
    for row in existing_rows.mappings().all():
        req_id = row["requirement_id"]
        layer = row["layer"] or "SYSTEM"
        content = _clean(row["content"])
        existing_keys.add((req_id, layer, content.lower()))
        key = (req_id, layer)
        next_versions[key] = max(next_versions.get(key, 0), int(row["version"] or 0))

    wb = _load_workbook()
    ws = wb["Interpretation Versioning"]
    rows = _sheet_rows(ws)

    inserted = 0
    skipped = 0

    for row in rows:
        req_code = _clean(row.get("Requirement ID"))
        layer = _normalize_layer(_clean(row.get("Layer")))
        content = _clean(row.get("Interpretation Text"))

        if not req_code or not layer or not content:
            skipped += 1
            continue

        requirement_id = requirement_by_code.get(req_code)
        if not requirement_id:
            skipped += 1
            continue

        dedupe_key = (str(requirement_id), layer, content.lower())
        if dedupe_key in existing_keys:
            skipped += 1
            continue

        version_key = (str(requirement_id), layer)
        version = next_versions.get(version_key, 0) + 1
        next_versions[version_key] = version

        await conn.execute(
            text(
                """
                INSERT INTO risk_interpretation (id, requirement_id, layer, content, version)
                VALUES (:id, :requirement_id, :layer, :content, :version)
                """
            ),
            {
                "id": _new_id(),
                "requirement_id": requirement_id,
                "layer": layer,
                "content": content,
                "version": version,
            },
        )

        existing_keys.add(dedupe_key)
        inserted += 1

    print(f"  Interpretations inserted: {inserted}, skipped: {skipped}")
    wb.close()
    return inserted, skipped


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def main() -> None:
    db_url = os.getenv("DATABASE_URL", DEFAULT_DB_URL)
    engine = create_async_engine(db_url, echo=False)

    n_controls = 0
    n_requirements = 0
    n_new_regs = 0
    n_metrics = 0
    n_ctrl_reqs = 0
    n_risks = 0
    n_interpretations = 0

    async with engine.begin() as conn:
        print("\n-- Sheet 1: Controls --")
        try:
            n_controls = await seed_controls(conn)
        except Exception as exc:
            print(f"  ERROR seeding controls: {exc}")

        print("\n-- Sheet 2: Requirements (Flat) + Regulations --")
        try:
            n_requirements, n_new_regs = await seed_requirements(conn)
        except Exception as exc:
            print(f"  ERROR seeding requirements/regulations: {exc}")

        print("\n-- Sheet 3: Control Metrics --")
        try:
            n_metrics = await seed_metric_definitions(conn)
        except Exception as exc:
            print(f"  ERROR seeding metric definitions: {exc}")

        print("\n-- Sheet 4: Requirements by Control links --")
        try:
            n_ctrl_reqs = await seed_control_requirements(conn)
        except Exception as exc:
            print(f"  ERROR seeding control-requirement links: {exc}")

        print("\n-- Baseline risk definitions --")
        try:
            n_risks = await seed_risk_definitions(conn)
        except Exception as exc:
            print(f"  ERROR seeding risk definitions: {exc}")

        print("\n-- Sheet 5: Interpretation Versioning --")
        try:
            n_interpretations, _ = await seed_interpretations(conn)
        except Exception as exc:
            print(f"  ERROR seeding interpretations: {exc}")

    await engine.dispose()

    print(
        "\nSeed summary: "
        f"controls={n_controls}, "
        f"requirements={n_requirements}, "
        f"new_regulations={n_new_regs}, "
        f"metric_definitions={n_metrics}, "
        f"control_requirement_links={n_ctrl_reqs}, "
        f"risk_definitions={n_risks}, "
        f"interpretations_inserted={n_interpretations}"
    )


if __name__ == "__main__":
    asyncio.run(main())


